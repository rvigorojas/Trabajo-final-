from io import BytesIO

from httpx import AsyncClient
from openpyxl import load_workbook

from app.models.usuario import Rol
from app.services.reporte_cierre import COLUMNAS_REPORTE_CIERRE
from tests.conftest import ahora_iso, auth_headers, crear_usuario_y_login


async def test_reporte_cierre_aeronautica_usa_columnas_reales_del_excel(client: AsyncClient):
    token = await crear_usuario_y_login(client, Rol.JEFE_RESCATE, "jefe.reporte1")
    # Sup. Gral. de Rescate convocado por la matriz real (Alerta II/III) para
    # poder verificar el mapeo M6/M3 -> nombre convocado.
    await crear_usuario_y_login(client, Rol.SUPERVISOR_GRAL_RESCATE, "sup.gral.reporte1")

    resp_activacion = await client.post(
        "/activaciones",
        json={
            "tipo_emergencia": "aeronautica",
            "nivel_alerta": "III",
            "tipo_alerta": 5,
            "tipo_incidente": "Fallo en el tren de aterrizaje",
            "hora_evento": "2026-08-15T10:30:00Z",
        },
        headers=auth_headers(token),
    )
    activacion_id = resp_activacion.json()["id"]

    await client.post(
        "/evaluaciones-iniciales",
        json={
            "activacion_id": activacion_id,
            "magnitud": "Aeronave con 90 pasajeros a bordo",
            "riesgos_secundarios": "Posible incendio en tren de aterrizaje",
            "hora_evento": ahora_iso(),
        },
        headers=auth_headers(token),
    )

    resp = await client.post(
        "/reportes-cierre", json={"activacion_id": activacion_id}, headers=auth_headers(token)
    )
    assert resp.status_code == 201, resp.text
    datos = resp.json()["datos"]

    # Mismas columnas, mismo nombre, que el Cuadro Estadístico real (28 en total).
    assert set(datos.keys()) == set(COLUMNAS_REPORTE_CIERRE["aeronautica"])
    assert datos["Fecha"] == "2026-08-15"
    assert datos["Mes"] == 8
    assert datos["Tipo de Alerta"] == 5
    assert datos["Problema"] == "Fallo en el tren de aterrizaje"
    assert "Posible incendio" in datos["Detalle / Observaciones"]
    assert datos["Supervisor Gral de Rescate M3"] == "Usuario sup.gral.reporte1"
    # Columnas fuera del alcance del PCE v1 (detalle operativo/administrativo
    # que el sistema no captura) quedan sin completar, no inventadas.
    assert datos["Compañía Aérea"] is None
    assert datos["Revisión\n(No llenar)"] is None


async def test_reporte_cierre_matpel_usa_clasificacion_real_sin_mapear_a_nivel(
    client: AsyncClient,
):
    token = await crear_usuario_y_login(client, Rol.DUTY_MANAGER, "duty.reporte2")

    resp_activacion = await client.post(
        "/activaciones",
        json={
            "tipo_emergencia": "matpel",
            "clasificacion_origen": "Clase 8",
            "tipo_incidente": "Derrame de ácido en plataforma",
            "hora_evento": "2026-08-15T14:00:00Z",
        },
        headers=auth_headers(token),
    )
    activacion_id = resp_activacion.json()["id"]

    resp = await client.post(
        "/reportes-cierre", json={"activacion_id": activacion_id}, headers=auth_headers(token)
    )
    assert resp.status_code == 201, resp.text
    datos = resp.json()["datos"]

    assert set(datos.keys()) == set(COLUMNAS_REPORTE_CIERRE["matpel"])
    # La clasificación UN se registra tal cual, sin mapeo a nivel_alerta
    # (PRD sección 8) — coincide con la columna real del Excel MATPEL.
    assert (
        datos['Clasificación MATPEL\n(Si es "otros" detallar en la columna siguiente)']
        == "Clase 8"
    )
    assert datos["Descripción de la emergencia"] == "Derrame de ácido en plataforma"


async def test_reporte_cierre_es_idempotente_por_activacion(client: AsyncClient):
    token = await crear_usuario_y_login(client, Rol.JEFE_RESCATE, "jefe.reporte3")
    resp_activacion = await client.post(
        "/activaciones",
        json={
            "tipo_emergencia": "aeronautica",
            "nivel_alerta": "II",
            "tipo_alerta": 2,
            "tipo_incidente": "Prueba idempotencia",
            "hora_evento": ahora_iso(),
        },
        headers=auth_headers(token),
    )
    activacion_id = resp_activacion.json()["id"]

    primera = await client.post(
        "/reportes-cierre", json={"activacion_id": activacion_id}, headers=auth_headers(token)
    )
    segunda = await client.post(
        "/reportes-cierre", json={"activacion_id": activacion_id}, headers=auth_headers(token)
    )
    assert primera.status_code == 201
    assert segunda.status_code == 201
    assert primera.json()["id"] == segunda.json()["id"]


async def test_exportar_reportes_cierre_arma_xlsx_con_columnas_reales(client: AsyncClient):
    token = await crear_usuario_y_login(client, Rol.DUTY_MANAGER, "duty.reporte4")
    resp_activacion = await client.post(
        "/activaciones",
        json={
            "tipo_emergencia": "matpel",
            "clasificacion_origen": "Clase 3",
            "tipo_incidente": "Derrame de combustible",
            "hora_evento": "2026-08-17T09:00:00Z",
        },
        headers=auth_headers(token),
    )
    activacion_id = resp_activacion.json()["id"]
    await client.post(
        "/reportes-cierre", json={"activacion_id": activacion_id}, headers=auth_headers(token)
    )

    resp = await client.get(
        "/reportes-cierre/exportar",
        params={"tipo_emergencia": "matpel"},
        headers=auth_headers(token),
    )
    assert resp.status_code == 200
    assert resp.headers["content-type"] == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert 'filename="cuadro-estadistico-matpel.xlsx"' in resp.headers["content-disposition"]

    libro = load_workbook(BytesIO(resp.content))
    hoja = libro["Base de Datos"]
    encabezado = [celda.value for celda in next(hoja.iter_rows(min_row=1, max_row=1))]
    assert encabezado == COLUMNAS_REPORTE_CIERRE["matpel"]

    columna_descripcion = encabezado.index("Descripción de la emergencia")
    filas = list(hoja.iter_rows(min_row=2, values_only=True))
    assert any(fila[columna_descripcion] == "Derrame de combustible" for fila in filas)
